import urllib.request
import time
import tkinter
import multiprocessing
import openpyxl

def ProbeInsertionWin(child_conn):
#GUI responsible for retrieving probe insertion time
	Insertion = tkinter.Tk()
	Insertion.geometry("300x150")
	Insertion.title("Probe Insertion")
	
	MinOption = tkinter.StringVar()
	SecOption = tkinter.StringVar()
	
	def WinKiller():
		child_conn.send((int(MinOption.get()) * 60) + int(SecOption.get()))
		Insertion.destroy()
	
	MinOption.set('1')
	SecOption.set('30')
	
	OptionsSec = ['00', '01', '02', '03', '04', '05', '06', '07', '08', '09',
				  '10', '11', '12', '13', '14', '15', '16', '17', '18', '19',
				  '20', '21', '22', '23', '24', '25', '26', '27', '28', '29',
				  '30', '31', '32', '33', '34', '35', '36', '37', '38', '39',
				  '40', '41', '42', '43', '44', '45', '46', '47', '48', '49',
				  '50', '51', '52', '53', '54', '55', '56', '57', '58', '59']
	OptionsMin = ['0', '1', '2', '3', '4', '5']
	
	frame1 = tkinter.Frame(Insertion)
	frame1.grid(column=0, row=1, sticky="nsew", padx=90, pady=7)

	label1 = tkinter.Label(Insertion, text="At what time was the probe inserted?")
	label2 = tkinter.Label(frame1, text=":")
	label1.grid(column=0, row=0, padx=15, pady=14)
	label2.grid(column=1, row=0)

	MinMenu = tkinter.OptionMenu(frame1, MinOption, *OptionsMin)
	SecMenu = tkinter.OptionMenu(frame1, SecOption, *OptionsSec)
	MinMenu.grid(column=0, row=0)
	SecMenu.grid(column=2, row=0)

	btn1 = tkinter.Button(Insertion, text="OK", command=WinKiller)
	btn1.grid(column=0, row=2, pady=14)
	
	Insertion.mainloop()
	
def TestReadyWin():
#GUI to start the timer
	ready = tkinter.Tk()
	ready.geometry("200x125")
	ready.title("Start Test")

	ok = tkinter.IntVar()

	def okClick():
		ok.set(1)
	def cancelClick():
		ok.set(0)
	
	but1 = tkinter.Button(ready, text='Start', command=okClick)
	but2 = tkinter.Button(ready, text='Cancel', command=cancelClick)
	
	but1.grid(column=0, row=0, padx=30, pady=30)
	but2.grid(column=1, row=0)
	
	but1.wait_variable(ok)

	ready.destroy()
	
	if(ok.get() == 1):
		return 1
	if(ok.get() == 0):
		return 0

def StartTestClick(pipeArray, eventArray):
#GUI for entering Test information
	NewTestInfo = [' ', ' ', ' ', ' ']  # [ProdName, Lot#, Test#/Letter, Probe#]
	ok = tkinter.IntVar()
	radBut = tkinter.IntVar()

	newTest = tkinter.Toplevel()
	newTest.geometry('350x125')
	newTest.title("New Test")
	lbl1 = tkinter.Label(newTest, text="Product Name")
	lbl2 = tkinter.Label(newTest, text="Lot #")
	lbl3 = tkinter.Label(newTest, text="Test #/Letter")
	lbl4 = tkinter.Label(newTest, text="Probe Number")
	lbl1.grid(column=0, row=0, sticky='E')
	lbl2.grid(column=0, row=1, sticky='E')
	lbl3.grid(column=0, row=2, sticky='E')
	lbl4.grid(column=0, row=3, sticky='E')

	txt1 = tkinter.Entry(newTest, width=15)
	txt2 = tkinter.Entry(newTest, width=15)
	txt3 = tkinter.Entry(newTest, width=5)
	txt1.grid(column=1, row=0)
	txt2.grid(column=1, row=1)
	txt3.grid(column=1, row=2, sticky='W')

	rad1 = tkinter.Radiobutton(newTest, text='1', value=1, variable=radBut)
	rad2 = tkinter.Radiobutton(newTest, text='2', value=2, variable=radBut)
	rad3 = tkinter.Radiobutton(newTest, text='3', value=3, variable=radBut)
	rad4 = tkinter.Radiobutton(newTest, text='4', value=4, variable=radBut)
	rad5 = tkinter.Radiobutton(newTest, text='5', value=5, variable=radBut)
	rad6 = tkinter.Radiobutton(newTest, text='6', value=6, variable=radBut)
	rad1.grid(column=1, row=3, sticky='E')
	rad2.grid(column=2, row=3)
	rad3.grid(column=3, row=3)
	rad4.grid(column=4, row=3)
	rad5.grid(column=5, row=3)
	rad6.grid(column=6, row=3)

	rad1.select()
	txt1.focus()
	
	Space1 = tkinter.Frame(newTest)
	Space1.grid(column=0, row=4)

	def okClick():
		ok.set(1)
	def cancelClick():
		ok.set(0)

	but1 = tkinter.Button(newTest, text='OK', width=10, command=okClick)
	but1.place(height=26, width=90, relx=0.15, rely=0.71)
	but2 = tkinter.Button(newTest, text='Cancel', width=10, command=cancelClick)
	but2.place(height=26, width=90, relx=0.50, rely=0.71)

	newTest.focus()
	txt1.focus()

	but1.wait_variable(ok)

	NewTestInfo[0] = txt1.get()
	NewTestInfo[1] = txt2.get()
	NewTestInfo[2] = txt3.get()
	NewTestInfo[3] = radBut.get()

	if(ok.get() == 1):
		p = multiprocessing.Process(target=TestStart, args=(NewTestInfo, pipeArray[NewTestInfo[3]-1], eventArray[NewTestInfo[3]-1]))	# -1 since NewTestInfo is an array that starts at [0] and probe numbers start at 1
		p.start()

	newTest.destroy()

def TestStart(NewTestInfo, child_conn, event):
# Starting point for recording test data
	fileName = fileNamer(NewTestInfo[0], NewTestInfo[1], NewTestInfo[2])

	parent_connProbeInsert, child_connProbeInsert = multiprocessing.Pipe()
	
	# initializes the excel workbook and sheet
	wb = openpyxl.Workbook()
	ws = wb.active

	Bold = openpyxl.styles.Font(bold=True, underline="single")

	ws['A1'] = 'TIME'
	ws['B1'] = 'TEMP'
	ws['D1'] = 'PRODUCT'
	ws['E1'] = 'LOT'
	ws['D2'] = NewTestInfo[0]
	ws['E2'] = NewTestInfo[1]
	ws['D4'] = 'FINAL WT'
	ws['D5'] = 'FINAL RT'
	ws['D6'] = 'FINAL EXO'
	ws['D7'] = 'TRUE RT'
	
	ws['A1'].font = Bold
	ws['B1'].font = Bold
	ws['D1'].font = Bold
	ws['E1'].font = Bold
	ws['D4'].font = Bold
	ws['D5'].font = Bold
	ws['D6'].font = Bold
	ws['D7'].font = Bold
	
	# After initializing the excel sheet, we start recording
	finalTimes = [0.0, 0.0, 0, 0.0]
	child_conn.send('testing')
	#child_conn.close()
	
	flag = TestReadyWin()
	
	q = multiprocessing.Process(target=ProbeInsertionWin, args=(child_connProbeInsert, ))
	q.start()
	
	if (flag == 1):
		finalTimes = timeKeeper(NewTestInfo[3], ws, child_conn, event)

		for i in (0, 1):
			finalTimes[i] = timeStandardizer(finalTimes[i])
		finalTimes[3] = timeStandardizer(finalTimes[3])
		# writes final times/temp on excel sheet
		ws['E4'] = finalTimes[0]
		ws['E5'] = finalTimes[1]
		ws['E6'] = finalTimes[2]
		ws['E7'] = finalTimes[3]
	probeTime = parent_connProbeInsert.recv()
	InsertionMarker(ws, probeTime)
	wb.save(fileName)
	child_conn.send('ready')
	child_conn.close()

	return

def fileNamer(prodName, prodLot, testNum):
	# assembles the string responsible for excel file name and location.
	fileName = r'\\EBSNAS\share\EManS\ProgTest\\' + prodName + '_' + prodLot + '_' + testNum + '.xlsx'
	return fileName

def timeKeeper(pNumber, ws, child_conn, event):
	# Counts Time for infoGrabber and Final Test Times
	flag = '1'
	trueRTflag = 1
	count = 2
	tempMax = int('0')
	finalTimes = [0.0, 0.0, 0, 0.0]  # [WT, RT, EXO, True RT]
	startTime = time.time()
	trueExo = [0.0, 0.0]

	while (flag == '1'):
		time.sleep(.75)  # stops the function for 0.9 seconds to ensure were not requesting temp more often then 1/sec
		curTime = time.time() - startTime
		curTemp = infoGrabber(pNumber)
		count = eXcelWriter(curTime, curTemp, count, ws)  # writes current sec count since start and current temp, function returns count + 1 to ensure next excel cell is used
		if (curTemp > tempMax + 100):
			child_conn.send('error')
			#child_conn.close()
		if (curTemp >= 170 and (curTemp > tempMax or curTemp + 1 == tempMax) and trueRTflag == 1):
			trueExo[0] = trueExo[1]
			trueExo[1] = curTime
		if (curTemp >= 170 and curTemp + 1 == tempMax):
			trueRTflag = 0
		if (curTemp > tempMax):  # updates max temp and time in an array
			tempMax = curTemp
			finalTimes[1] = time.time() - startTime
			finalTimes[2] = tempMax
		if (curTemp >= 105 and finalTimes[0] == 0.0):  # records WT
			finalTimes[0] = time.time() - startTime
			child_conn.send('hot')
			#child_conn.close()
			#print("105F TIMES MET/RECORDED!")  # Only used during testing

		if ((tempMax - 10) > curTemp and curTemp < tempMax):  # breaks the while loop to end the function
			flag = '0'
		if (event.is_set()):
			event.clear()
			flag = '0'
	finalTimes[3] = (trueExo[0] + trueExo[1]) / 2
	return (finalTimes)

def infoGrabber(pNumber):
	# Requests current temperature information from Yokogawa DX100P and returns the requested probe temp
	with urllib.request.urlopen('http://192.168.1.137/cgi-bin/ope/allch.cgi') as f:
		raw_data = f.read()  # converts source information from yokogawa into a string
	if (pNumber == 1):  # based on probe number, cuts all information except requested probe temp from the string
		raw_data = raw_data[2269:-4035]
		raw_data = int(raw_data.decode('utf-8'))
		return raw_data
	elif (pNumber == 2):
		raw_data = raw_data[3056:-3248]
		raw_data = int(raw_data.decode('utf-8'))
		return raw_data
	elif (pNumber == 3):
		raw_data = raw_data[3842:-2461]
		raw_data = int(raw_data.decode('utf-8'))
		return raw_data
	elif (pNumber == 4):
		raw_data = raw_data[4629:-1674]
		raw_data = int(raw_data.decode('utf-8'))
		return raw_data
	elif (pNumber == 5):
		raw_data = raw_data[5416:-887]
		raw_data = int(raw_data.decode('utf-8'))
		return raw_data
	elif (pNumber == 6):
		raw_data = raw_data[6203:-100]
		raw_data = int(raw_data.decode('utf-8'))
		return raw_data

def eXcelWriter(curTime, curTemp, count, ws):
	# Writes read data from Yokogawa to excel sheet
	ws.cell(row=count, column=1, value=curTime)
	ws.cell(row=count, column=2, value=curTemp)
	#print("Writing to Excel File")  # Only used in testing
	return count + 1

def timeStandardizer(time):
	# Converts from decimal minutes to min:sec
	x = 0.0
	time = (time - (time % 1))
	x = time / 60
	x = (x - (x % 1))
	if (int(time - (x * 60))) < 10:
		time = str(int(x)) + ':' + '0' + str(int(time - (x * 60)))	
	else:
		time = str(int(x)) + ':' + str(int(time - (x * 60)))
	return time

def InsertionMarker(ws, probeTime):
	#searches for the correct time of probe insertion and highlights those cells
	flag = 0

	my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
	my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)

	for row in ws.iter_rows(min_row=2, min_col=1, max_col=1, max_row=400):
		for cell in row:
			if(flag == 0):
				if ((cell.value - (cell.value % 1)) == probeTime):
					ws[cell.coordinate].fill = my_fill
					flag = 1
	return